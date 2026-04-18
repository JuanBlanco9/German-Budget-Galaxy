#!/usr/bin/env python3
"""
Parse gov.scot 'Council Tax by Band' XLSX into the same structure as the
English MHCLG council tax JSON so the frontend lookup works uniformly.

Source: https://www.gov.scot/publications/council-tax-datasets/
Output: data/uk/fiscal/council_tax/scotland_council_tax_2024_25.json
"""
import hashlib
import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'data' / 'uk' / 'fiscal' / 'council_tax' / 'scotland_ct_by_band_2024_25.xlsx'
OUT = ROOT / 'data' / 'uk' / 'fiscal' / 'council_tax' / 'scotland_council_tax_2024_25.json'


def sha256(p: Path) -> str:
    h = hashlib.sha256()
    with open(p, 'rb') as f:
        for c in iter(lambda: f.read(1 << 20), b''):
            h.update(c)
    return h.hexdigest()


def main():
    if not XLSX.exists():
        print(f'ERROR: {XLSX} not found', file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # Header is at row index 4: ['', 'Band A', 'Band B', ..., 'Band H']
    # Council data starts at row 7.
    councils = []
    for row in rows[7:]:
        if not row or not isinstance(row[0], str):
            continue
        name = row[0].strip()
        if not name or name.lower().startswith('scotland'):
            continue

        bands = {}
        for i, band in enumerate('ABCDEFGH'):
            v = row[1 + i]
            if isinstance(v, (int, float)):
                bands[f'band_{band}'] = round(float(v), 2)
        if not bands:
            continue

        councils.append({
            'e_code': None,
            'ons_code': None,
            'name': name,
            'region_code': 'SC',
            'class_code': 'Scottish LA',
            'area_type': 'Scotland',
            **bands,
        })

    out = {
        'country': 'uk',
        'subject': 'council_tax_bands',
        'jurisdiction': 'scotland',
        'fiscal_year_label': '2024-25',
        'fiscal_year_start': 2024,
        'unit': 'GBP_per_year',
        'note': (
            'Scottish council tax by band, 2024-25. All 32 Scottish councils. '
            'Band D figure is the statutory council tax; A/B/C/E/F/G/H derived '
            'via the statutory ratios (240/280/320/473/585/705/882 of 360 for Band D). '
            'Argyll & Bute was the only council that froze rates in 2024-25.'
        ),
        'source': {
            'publisher': 'Scottish Government',
            'dataset': 'Council Tax by band, 2024-25',
            'landing_url': 'https://www.gov.scot/publications/council-tax-datasets/',
            'file_path': str(XLSX.relative_to(ROOT)).replace('\\', '/'),
            'file_sha256': sha256(XLSX),
            'downloaded_at': date.today().isoformat(),
        },
        'council_count': len(councils),
        'councils': councils,
    }

    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    print(f'Wrote {OUT.relative_to(ROOT)}')
    print(f'  {len(councils)} Scottish councils')
    for c in councils[:4]:
        print(f'    {c["name"]:30s} Band D £{c["band_D"]:>8,.2f}')


if __name__ == '__main__':
    main()
