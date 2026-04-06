#!/usr/bin/env python3
"""
Build Budget Galaxy tree JSON files for German Laender and Kommunen spending.

Data sources:
  - Destatis Rechnungsergebnis Excel files (2019, 2020, 2021):
    Sheet "71711-12" has Laender+Gemeinden expenditure by Aufgabenbereiche and state.
    Values in Millionen EUR. We convert to EUR integers.
  - Eurostat COFOG gov_10a_exp API (S1312=State, S1313=Local):
    10 COFOG categories, 2015-2023. Values in MIO_EUR. Used as fallback for
    years without Destatis Excel (2015-2018, 2022-2023).

Output:
  data/de/laender/laender_parsed_YYYY.json  -- State-level tree with 16 Laender x ~15 functions
  data/de/laender/kommunen_parsed_YYYY.json -- Kommunen aggregate with 10 COFOG categories
  data/de/laender/laender_eurostat_YYYY.json -- Eurostat-based Laender aggregate (fallback years)

Author: build_laender_trees.py for Budget Galaxy
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data", "de", "laender")

# Column mapping for sheet 71711-12 (same across 2019-2021)
# Col 3 = Laender zusammen, Cols 4-19 = individual states
STATE_COLUMNS = {
    4:  ("bw",  "Baden-Württemberg"),
    5:  ("by",  "Bayern"),
    6:  ("bb",  "Brandenburg"),
    7:  ("he",  "Hessen"),
    8:  ("mv",  "Mecklenburg-Vorpommern"),
    9:  ("ni",  "Niedersachsen"),
    10: ("nw",  "Nordrhein-Westfalen"),
    11: ("rp",  "Rheinland-Pfalz"),
    12: ("sl",  "Saarland"),
    13: ("sn",  "Sachsen"),
    14: ("st",  "Sachsen-Anhalt"),
    15: ("sh",  "Schleswig-Holstein"),
    16: ("th",  "Thüringen"),
    17: ("be",  "Berlin"),
    18: ("hb",  "Bremen"),
    19: ("hh",  "Hamburg"),
}

# Top-level Aufgabenbereiche hierarchy
# Each entry: (code, id_suffix, german_name, [sub_codes])
# Sub-codes are children shown under the parent; they are NOT separate top-level items.
# "Insgesamt" (1001) is the root total.
# "Versorgung" (1060) is cross-cutting (pensions for civil servants across functions).

TOP_LEVEL_FUNCTIONS = [
    ("1002", "verwaltung",    "Politische Führung und zentrale Verwaltung",
        [("1003", "pol_fuehrung",  "Politische Führung"),
         ("1004", "innere_verw",   "Innere Verwaltung")]),
    ("1005", "auswaertiges",  "Auswärtige Angelegenheiten",
        [("1006", "entw_zusammenarbeit", "Wirtschaftliche Zusammenarbeit und Entwicklung")]),
    ("1007", "verteidigung",  "Verteidigung", []),
    ("1008", "sicherheit",    "Öffentliche Sicherheit und Ordnung",
        [("1009", "polizei", "Polizei")]),
    ("1010", "rechtsschutz",  "Rechtsschutz", []),
    ("1011", "finanzverwaltung", "Finanzverwaltung", []),
    ("1012", "schulen",       "Allgemeinbildende und berufliche Schulen",
        [("1013", "allg_schulen",  "Allgemeinbildende Schulen"),
         ("1014", "beruf_schulen", "Berufliche Schulen")]),
    ("1015", "hochschulen",   "Hochschulen", []),
    ("1016", "foerderung_bildung", "Förderung für Schüler, Studierende u. dgl.", []),
    ("1017", "sonst_bildung", "Sonstiges Bildungswesen", []),
    ("1018", "wissenschaft",  "Wissenschaft, Forschung, Entwicklung", []),
    ("1019", "kultur",        "Kultur und Religion",
        [("1020", "theater_musik", "Theater und Musik")]),
    ("1021", "soziales",      "Soziale Sicherung, Familie und Jugend",
        [("1022", "soz_verwaltung",     "Verwaltung für soziale Angelegenheiten"),
         ("1023", "sozialversicherung", "Sozialversicherung"),
         ("1024", "familienhilfe",      "Familienhilfe, Wohlfahrtspflege"),
         ("1025", "kriegsfolgen",       "Soziale Leistungen für Kriegsfolgen"),
         ("1026", "arbeitsmarkt",       "Arbeitsmarktpolitik"),
         ("1027", "jugendhilfe",        "Kinder- und Jugendhilfe (ohne Kita)"),
         ("1028", "kita",               "Kindertagesbetreuung"),
         ("1029", "sgb_xii",            "Sozialleistungen SGB XII/AsylbLG/Eingliederung")]),
    ("1030", "gesundheit",    "Gesundheit, Umwelt, Sport und Erholung",
        [("1031", "gesundheitswesen", "Gesundheitswesen"),
         ("1032", "sport_erholung",   "Sport und Erholung"),
         ("1033", "umweltschutz",     "Umwelt- und Naturschutz")]),
    ("1034", "wohnungswesen", "Wohnungswesen, Städtebau, Raumordnung",
        [("1035", "wohnungsbau",      "Wohnungswesen, Wohnungsbauprämie"),
         ("1036", "raumordnung",      "Raumordnung, Städtebauförderung"),
         ("1037", "gemeinschaftsdienste", "Kommunale Gemeinschaftsdienste")]),
    ("1038", "landwirtschaft", "Ernährung, Landwirtschaft und Forsten", []),
    ("1039", "gewerbe",       "Energie- und Wasserwirtschaft, Gewerbe, Dienstleistungen",
        [("1040", "energie_entsorgung",  "Energie, Wasserversorgung, Entsorgung"),
         ("1041", "abwasser",            "Abwasserentsorgung"),
         ("1042", "abfall",              "Abfallwirtschaft"),
         ("1043", "versicherungswesen",  "Geld- und Versicherungswesen"),
         ("1044", "sonst_gewerbe",       "Sonstiges Gewerbe und Dienstleistungen"),
         ("1045", "reg_foerderung",      "Regionale Fördermaßnahmen")]),
    ("1046", "verkehr",       "Verkehrs- und Nachrichtenwesen",
        [("1047", "strassen",            "Straßen einschl. Verwaltung"),
         ("1048", "autobahnen",          "Bundesautobahnen"),
         ("1049", "bundes_landesstr",    "Bundes- und Landesstraßen"),
         ("1050", "kreisstrassen",       "Kreisstraßen"),
         ("1051", "gemeindestrassen",    "Gemeindestraßen"),
         ("1052", "wasserstrassen",      "Wasserstraßen und Häfen"),
         ("1053", "eisenbahn_oepnv",     "Eisenbahnen und ÖPNV"),
         ("1054", "sonst_verkehr",       "Sonstiges Verkehrswesen")]),
    ("1055", "finanzwirtschaft", "Finanzwirtschaft",
        [("1056", "kapitalvermoegen",    "Grund- und Kapitalvermögen"),
         ("1057", "steuern_finzuw",      "Steuern und Finanzzuweisungen"),
         ("1058", "schulden",            "Schulden"),
         ("1059", "beihilfen",           "Beihilfen und Unterstützungen")]),
    ("1060", "versorgung",    "Versorgung (Beamtenpensionen)", []),
]

# COFOG category mapping for Eurostat data
COFOG_LABELS = {
    "TOTAL": "Total",
    "GF01": "General public services",
    "GF02": "Defence",
    "GF03": "Public order and safety",
    "GF04": "Economic affairs",
    "GF05": "Environmental protection",
    "GF06": "Housing and community amenities",
    "GF07": "Health",
    "GF08": "Recreation, culture and religion",
    "GF09": "Education",
    "GF10": "Social protection",
}

COFOG_IDS = {
    "GF01": "allg_dienste",
    "GF02": "verteidigung",
    "GF03": "sicherheit",
    "GF04": "wirtschaft",
    "GF05": "umwelt",
    "GF06": "wohnungswesen",
    "GF07": "gesundheit",
    "GF08": "kultur_freizeit",
    "GF09": "bildung",
    "GF10": "soziales",
}

COFOG_GERMAN = {
    "GF01": "Allgemeine öffentliche Verwaltung",
    "GF02": "Verteidigung",
    "GF03": "Öffentliche Ordnung und Sicherheit",
    "GF04": "Wirtschaftliche Angelegenheiten",
    "GF05": "Umweltschutz",
    "GF06": "Wohnungswesen und kommunale Einrichtungen",
    "GF07": "Gesundheitswesen",
    "GF08": "Freizeit, Sport, Kultur und Religion",
    "GF09": "Bildungswesen",
    "GF10": "Soziale Sicherung",
}

# Destatis Excel files
DESTATIS_FILES = {
    2019: "rechnungsergebnis_2019.xlsx",
    2020: "rechnungsergebnis_2020.xlsx",
    2021: "rechnungsergebnis_2021.xlsx",
}

# Data start rows per year (where code 1001 begins)
DESTATIS_START_ROWS = {
    2019: 7,
    2020: 6,
    2021: 6,
}

# Rows per Aufgabenbereich block
BLOCK_SIZE = 63

# Within each block, offset to key rows (0-indexed from header):
# .43 = Bereinigte Ausgaben (total L+G)    -> offset 42
# .44 = Länder                              -> offset 43
# .45 = Gemeinden/Gv.                       -> offset 44
OFFSET_BEREINIGTE_TOTAL = 42
OFFSET_BEREINIGTE_LAENDER = 43
OFFSET_BEREINIGTE_GEMEINDEN = 44


# ---------------------------------------------------------------------------
# Destatis Excel Parser
# ---------------------------------------------------------------------------

def parse_value(v):
    """Convert Excel cell value to integer EUR (from Millionen EUR).
    Returns 0 for dashes, None, or non-numeric values."""
    if v is None:
        return 0
    if isinstance(v, str):
        v = v.strip()
        if v in ('-', '.', '', '–', '...', 'x', 'X'):
            return 0
        try:
            v = float(v.replace(',', '.'))
        except ValueError:
            return 0
    try:
        # Values in Millionen EUR -> convert to EUR
        return int(round(float(v) * 1_000_000))
    except (ValueError, TypeError):
        return 0


def find_aufgabenbereich_rows(ws, start_row):
    """Scan the worksheet to find all Aufgabenbereich header rows.
    Returns dict: code -> row_number"""
    result = {}
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=2, values_only=False):
        r = row[0].row
        a = str(row[0].value) if row[0].value else ''
        if '\n' in a:
            parts = a.split('\n')
            first = parts[0].strip()
            if first and '.' not in first and first.isdigit():
                result[first] = r
    return result


def get_row_values(ws, row_num, cols):
    """Get values from a specific row for given columns.
    Returns dict: col -> raw_value"""
    result = {}
    for row in ws.iter_rows(min_row=row_num, max_row=row_num,
                            min_col=min(cols), max_col=max(cols),
                            values_only=False):
        for c in row:
            if c.column in cols:
                result[c.column] = c.value
    return result


def extract_laender_data(ws, aufgaben_rows, code, level="laender"):
    """Extract per-state values for a given Aufgabenbereich code.
    level: 'laender' for state spending, 'gemeinden' for municipal spending.
    Returns dict: state_abbrev -> value_in_EUR, plus 'total' key."""
    if code not in aufgaben_rows:
        return None

    header_row = aufgaben_rows[code]
    if level == "laender":
        data_row = header_row + OFFSET_BEREINIGTE_LAENDER
    elif level == "gemeinden":
        data_row = header_row + OFFSET_BEREINIGTE_GEMEINDEN
    else:
        data_row = header_row + OFFSET_BEREINIGTE_TOTAL

    cols_needed = [3] + list(STATE_COLUMNS.keys())
    raw = get_row_values(ws, data_row, cols_needed)

    result = {"total": parse_value(raw.get(3))}
    for col, (abbrev, _name) in STATE_COLUMNS.items():
        result[abbrev] = parse_value(raw.get(col))

    return result


def build_function_children(ws, aufgaben_rows, parent_code, sub_codes,
                            state_abbrev, level="laender"):
    """Build children array for a function within one state.
    Returns list of {id, name, value} dicts, sorted by value desc."""
    children = []
    for sub_code, sub_id, sub_name in sub_codes:
        data = extract_laender_data(ws, aufgaben_rows, sub_code, level)
        if data is None:
            continue
        val = data.get(state_abbrev, 0)
        if val > 0:
            children.append({
                "id": f"laender_{state_abbrev}_{sub_id}",
                "name": sub_name,
                "value": val
            })
    children.sort(key=lambda x: x["value"], reverse=True)
    return children


def build_state_tree(ws, aufgaben_rows, state_abbrev, state_name):
    """Build the function-level tree for one state.
    Returns {id, name, value, children} for the state."""
    # Get total for the state (1001 = Insgesamt)
    total_data = extract_laender_data(ws, aufgaben_rows, "1001", "laender")
    state_total = total_data.get(state_abbrev, 0) if total_data else 0

    func_children = []
    for code, func_id, func_name, sub_codes in TOP_LEVEL_FUNCTIONS:
        data = extract_laender_data(ws, aufgaben_rows, code, "laender")
        if data is None:
            continue
        val = data.get(state_abbrev, 0)
        if val <= 0:
            continue

        node = {
            "id": f"laender_{state_abbrev}_{func_id}",
            "name": func_name,
            "value": val,
        }

        # Add sub-function children if available
        sub_children = build_function_children(
            ws, aufgaben_rows, code, sub_codes, state_abbrev, "laender"
        )
        if sub_children:
            node["children"] = sub_children

        func_children.append(node)

    func_children.sort(key=lambda x: x["value"], reverse=True)

    return {
        "id": f"laender_{state_abbrev}",
        "name": state_name,
        "value": state_total,
        "children": func_children,
    }


def build_laender_tree_from_destatis(year):
    """Build the full Laender tree from a Destatis Excel file."""
    fname = DESTATIS_FILES.get(year)
    if not fname:
        print(f"  No Destatis file for {year}")
        return None

    fpath = os.path.join(DATA_DIR, fname)
    if not os.path.exists(fpath):
        print(f"  File not found: {fpath}")
        return None

    print(f"  Loading {fname}...")
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

    ws = wb['71711-12']
    start_row = DESTATIS_START_ROWS.get(year, 6)

    print(f"  Finding Aufgabenbereiche (start_row={start_row})...")
    aufgaben_rows = find_aufgabenbereich_rows(ws, start_row)
    print(f"  Found {len(aufgaben_rows)} Aufgabenbereiche: {sorted(aufgaben_rows.keys())[:10]}...")

    # Get overall total
    total_data = extract_laender_data(ws, aufgaben_rows, "1001", "laender")
    overall_total = total_data.get("total", 0) if total_data else 0

    print(f"  Building state trees (overall Laender total: {overall_total:,} EUR)...")

    state_children = []
    for col in sorted(STATE_COLUMNS.keys()):
        abbrev, name = STATE_COLUMNS[col]
        state_tree = build_state_tree(ws, aufgaben_rows, abbrev, name)
        if state_tree["value"] > 0:
            state_children.append(state_tree)

    state_children.sort(key=lambda x: x["value"], reverse=True)

    tree = {
        "id": "laender",
        "name": "Länder (Bundesländer)",
        "value": overall_total,
        "year": year,
        "source": f"Destatis Rechnungsergebnis {year}, Tabelle 71711-12, Bereinigte Ausgaben der Länder",
        "unit": "EUR",
        "children": state_children,
    }

    wb.close()
    return tree


# ---------------------------------------------------------------------------
# Eurostat COFOG Parser
# ---------------------------------------------------------------------------

def parse_eurostat_json(fpath):
    """Parse a Eurostat JSON-stat response.
    Returns dict: {year: {cofog_code: value_MIO_EUR}}"""
    with open(fpath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Dimensions: freq(1) x unit(1) x sector(1) x cofog99(11) x na_item(1) x geo(1) x time(N)
    sizes = data["size"]  # [1, 1, 1, 11, 1, 1, 9]
    n_cofog = sizes[3]
    n_time = sizes[6]

    cofog_index = data["dimension"]["cofog99"]["category"]["index"]  # {code: idx}
    time_index = data["dimension"]["cofog99"]["category"]["index"]
    time_labels = data["dimension"]["time"]["category"]["index"]  # {year: idx}

    values = data["value"]

    result = {}
    for year_str, t_idx in time_labels.items():
        year = int(year_str)
        result[year] = {}
        for cofog_code, c_idx in cofog_index.items():
            # Flat index: cofog_idx * n_time + time_idx
            flat_idx = c_idx * n_time + t_idx
            val = values.get(str(flat_idx))
            if val is not None:
                result[year][cofog_code] = val  # MIO_EUR
            else:
                result[year][cofog_code] = 0.0

    return result


def build_eurostat_laender_tree(year, eurostat_data):
    """Build a Laender aggregate tree from Eurostat S1312 data for one year.
    This has only 10 COFOG categories (no per-state breakdown)."""
    if year not in eurostat_data:
        return None

    year_data = eurostat_data[year]
    total_mio = year_data.get("TOTAL", 0)
    total_eur = int(round(total_mio * 1_000_000))

    children = []
    for cofog_code in ["GF01", "GF02", "GF03", "GF04", "GF05",
                        "GF06", "GF07", "GF08", "GF09", "GF10"]:
        val_mio = year_data.get(cofog_code, 0)
        val_eur = int(round(val_mio * 1_000_000))
        if val_eur > 0:
            children.append({
                "id": f"laender_{COFOG_IDS[cofog_code]}",
                "name": COFOG_GERMAN[cofog_code],
                "value": val_eur,
            })

    children.sort(key=lambda x: x["value"], reverse=True)

    return {
        "id": "laender",
        "name": "Länder (Bundesländer)",
        "value": total_eur,
        "year": year,
        "source": f"Eurostat gov_10a_exp, sector S1312 (State government), {year}",
        "unit": "EUR",
        "note": "Aggregate only (no per-state breakdown). 10 COFOG Level-1 categories.",
        "children": children,
    }


def build_kommunen_tree(year, eurostat_data):
    """Build a Kommunen tree from Eurostat S1313 data for one year."""
    if year not in eurostat_data:
        return None

    year_data = eurostat_data[year]
    total_mio = year_data.get("TOTAL", 0)
    total_eur = int(round(total_mio * 1_000_000))

    children = []
    for cofog_code in ["GF01", "GF02", "GF03", "GF04", "GF05",
                        "GF06", "GF07", "GF08", "GF09", "GF10"]:
        val_mio = year_data.get(cofog_code, 0)
        val_eur = int(round(val_mio * 1_000_000))
        if val_eur > 0:
            children.append({
                "id": f"kommunen_{COFOG_IDS[cofog_code]}",
                "name": COFOG_GERMAN[cofog_code],
                "value": val_eur,
            })

    children.sort(key=lambda x: x["value"], reverse=True)

    return {
        "id": "kommunen",
        "name": "Kommunen (Gemeinden)",
        "value": total_eur,
        "year": year,
        "source": f"Eurostat gov_10a_exp, sector S1313 (Local government), {year}",
        "unit": "EUR",
        "children": children,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def save_json(data, filepath):
    """Save tree as formatted JSON."""
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    size_kb = os.path.getsize(filepath) / 1024
    print(f"  Saved: {filepath} ({size_kb:.1f} KB)")


def main():
    print("=" * 70)
    print("Building German Laender & Kommunen budget trees")
    print("=" * 70)

    # --- Parse Eurostat data ---
    print("\n--- Parsing Eurostat S1312 (Laender) ---")
    s1312_path = os.path.join(DATA_DIR, "eurostat_s1312_laender.json")
    s1313_path = os.path.join(DATA_DIR, "eurostat_s1313_kommunen.json")

    eurostat_laender = {}
    eurostat_kommunen = {}

    if os.path.exists(s1312_path):
        eurostat_laender = parse_eurostat_json(s1312_path)
        for y in sorted(eurostat_laender.keys()):
            total = eurostat_laender[y].get("TOTAL", 0)
            print(f"  {y}: Total = {total:,.0f} MIO EUR")
    else:
        print(f"  WARNING: {s1312_path} not found")

    print("\n--- Parsing Eurostat S1313 (Kommunen) ---")
    if os.path.exists(s1313_path):
        eurostat_kommunen = parse_eurostat_json(s1313_path)
        for y in sorted(eurostat_kommunen.keys()):
            total = eurostat_kommunen[y].get("TOTAL", 0)
            print(f"  {y}: Total = {total:,.0f} MIO EUR")
    else:
        print(f"  WARNING: {s1313_path} not found")

    # --- Build Destatis-based trees (2019-2021) ---
    destatis_years = sorted(DESTATIS_FILES.keys())

    for year in destatis_years:
        print(f"\n--- Building Laender tree from Destatis {year} ---")
        tree = build_laender_tree_from_destatis(year)
        if tree:
            outpath = os.path.join(DATA_DIR, f"laender_parsed_{year}.json")
            save_json(tree, outpath)

            # Summary
            n_states = len(tree["children"])
            top3 = [(c["name"], c["value"]) for c in tree["children"][:3]]
            print(f"  {n_states} states. Top 3: {', '.join(f'{n} ({v:,} EUR)' for n, v in top3)}")

    # --- Build Kommunen trees from Eurostat for all years ---
    print(f"\n--- Building Kommunen trees from Eurostat (all years) ---")
    for year in sorted(eurostat_kommunen.keys()):
        tree = build_kommunen_tree(year, eurostat_kommunen)
        if tree:
            outpath = os.path.join(DATA_DIR, f"kommunen_parsed_{year}.json")
            save_json(tree, outpath)

    # --- Build Eurostat Laender fallback trees for years without Destatis ---
    eurostat_only_years = [y for y in sorted(eurostat_laender.keys())
                           if y not in destatis_years]
    if eurostat_only_years:
        print(f"\n--- Building Eurostat Laender fallback trees ({eurostat_only_years}) ---")
        for year in eurostat_only_years:
            tree = build_eurostat_laender_tree(year, eurostat_laender)
            if tree:
                outpath = os.path.join(DATA_DIR, f"laender_eurostat_{year}.json")
                save_json(tree, outpath)

    # --- Also save Eurostat Laender trees for Destatis years (cross-reference) ---
    print(f"\n--- Building Eurostat Laender trees for Destatis years (cross-reference) ---")
    for year in destatis_years:
        if year in eurostat_laender:
            tree = build_eurostat_laender_tree(year, eurostat_laender)
            if tree:
                outpath = os.path.join(DATA_DIR, f"laender_eurostat_{year}.json")
                save_json(tree, outpath)

    print("\n" + "=" * 70)
    print("Source documentation:")
    print("  Destatis years (2019-2021): Full 16-state x ~20 function breakdown")
    print("    laender_parsed_YYYY.json from Rechnungsergebnis Excel, Tabelle 71711-12")
    print("  Eurostat years (2015-2023): 10 COFOG categories, aggregate only")
    print("    laender_eurostat_YYYY.json from gov_10a_exp API, sector S1312")
    print("    kommunen_parsed_YYYY.json from gov_10a_exp API, sector S1313")
    print("  All values in EUR integers.")
    print("=" * 70)


if __name__ == "__main__":
    main()
