#!/usr/bin/env python3
"""
Convert Warwickshire's 12 monthly XLSX (sheet 'pound_500') to CSV.
Source: api.warwickshire.gov.uk/documents/WCCC-428063900-{ID}
"""
import csv
from pathlib import Path
import openpyxl

ID_MAP = {
    "2024_04_april":     1929,
    "2024_05_may":       1931,
    "2024_06_june":      1933,
    "2024_07_july":      1952,
    "2024_08_august":    1974,
    "2024_09_september": 1976,
    "2024_10_october":   1992,
    "2024_11_november":  1991,
    "2024_12_december":  2028,
    "2025_01_january":   2031,
    "2025_02_february":  2047,
    "2025_03_march":     2048,
}

CACHE = Path(r"C:\tmp")
DEST = Path(r"D:\germany-ngo-map\data\uk\local_authorities\spend\warwickshire_county_council")
DEST.mkdir(parents=True, exist_ok=True)


def convert(src: Path, dest: Path):
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    if "pound_500" in wb.sheetnames:
        ws = wb["pound_500"]
    else:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # Find header row (first row with 'SuppID' as first column or any cell)
    header_idx = 0
    for i, r in enumerate(rows[:10]):
        if r and any(c == "SuppID" for c in r if c is not None):
            header_idx = i
            break
    rows = rows[header_idx:]
    # Trim trailing empty cols
    if rows:
        keep_cols = max(i + 1 for i, c in enumerate(rows[0]) if c not in (None, ""))
        rows = [r[:keep_cols] for r in rows]
    with open(dest, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        for row in rows:
            w.writerow(["" if v is None else v for v in row])


total = 0
for slug, doc_id in ID_MAP.items():
    src = CACHE / f"wk_{doc_id}.xlsx"
    if not src.exists():
        print(f"  MISSING: {src.name}")
        continue
    dest = DEST / f"payments_{slug}.csv"
    convert(src, dest)
    n = sum(1 for _ in open(dest, encoding="utf-8-sig")) - 1
    print(f"  {slug}: {n:>6} rows  {dest.name}")
    total += n
print(f"\nTotal: {total} rows")
