#!/usr/bin/env python3
"""
Parse Destatis Rechnungsergebnis Excel files for 2017 and 2018.
Generates laender_parsed_YYYY.json in the same format as 2019-2021.

2017 uses Fachserie format (sheet "2.2"), 2018 uses Statistischer Bericht (sheet "71711-12").
"""

import json
import os
import sys
import time

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data", "de", "laender")

STATE_COLUMNS = {
    4:  ("bw",  "Baden-Württemberg"),
    5:  ("by",  "Bayern"),
    6:  ("bb",  "Brandenburg"),
    7:  ("he",  "Hessen"),
    8:  ("mv",  "Mecklenburg-Vorpommern"),
    9:  ("ni",  "Niedersachsen"),
    10: ("nw",  "Nordrhein-Westfalen"),
    11: ("rp",  "Rheinland-Pfalz"),
    12: ("sl",  "Saarland"),
    13: ("sn",  "Sachsen"),
    14: ("st",  "Sachsen-Anhalt"),
    15: ("sh",  "Schleswig-Holstein"),
    16: ("th",  "Thüringen"),
    17: ("be",  "Berlin"),
    18: ("hb",  "Bremen"),
    19: ("hh",  "Hamburg"),
}

TOP_LEVEL_FUNCTIONS = [
    ("1002", "verwaltung",    "Politische Führung und zentrale Verwaltung",
        [("1003", "pol_fuehrung",  "Politische Führung"),
         ("1004", "innere_verw",   "Innere Verwaltung")]),
    ("1005", "auswaertiges",  "Auswärtige Angelegenheiten",
        [("1006", "entw_zusammenarbeit", "Wirtschaftliche Zusammenarbeit und Entwicklung")]),
    ("1007", "verteidigung",  "Verteidigung", []),
    ("1008", "sicherheit",    "Öffentliche Sicherheit und Ordnung",
        [("1009", "polizei", "Polizei")]),
    ("1010", "rechtsschutz",  "Rechtsschutz", []),
    ("1011", "finanzverwaltung", "Finanzverwaltung", []),
    ("1012", "schulen",       "Allgemeinbildende und berufliche Schulen",
        [("1013", "allg_schulen",  "Allgemeinbildende Schulen"),
         ("1014", "beruf_schulen", "Berufliche Schulen")]),
    ("1015", "hochschulen",   "Hochschulen", []),
    ("1016", "foerderung_bildung", "Förderung für Schüler, Studierende u. dgl.", []),
    ("1017", "sonst_bildung", "Sonstiges Bildungswesen", []),
    ("1018", "wissenschaft",  "Wissenschaft, Forschung, Entwicklung", []),
    ("1019", "kultur",        "Kultur und Religion",
        [("1020", "theater_musik", "Theater und Musik")]),
    ("1021", "soziales",      "Soziale Sicherung, Familie und Jugend",
        [("1022", "soz_verwaltung",     "Verwaltung für soziale Angelegenheiten"),
         ("1023", "sozialversicherung", "Sozialversicherung"),
         ("1024", "familienhilfe",      "Familienhilfe, Wohlfahrtspflege"),
         ("1025", "kriegsfolgen",       "Soziale Leistungen für Kriegsfolgen"),
         ("1026", "arbeitsmarkt",       "Arbeitsmarktpolitik"),
         ("1027", "jugendhilfe",        "Kinder- und Jugendhilfe (ohne Kita)"),
         ("1028", "kita",               "Kindertagesbetreuung"),
         ("1029", "sgb_xii",            "Sozialleistungen SGB XII/AsylbLG/Eingliederung")]),
    ("1030", "gesundheit",    "Gesundheit, Umwelt, Sport und Erholung",
        [("1031", "gesundheitswesen", "Gesundheitswesen"),
         ("1032", "sport_erholung",   "Sport und Erholung"),
         ("1033", "umweltschutz",     "Umwelt- und Naturschutz")]),
    ("1034", "wohnungswesen", "Wohnungswesen, Städtebau, Raumordnung",
        [("1035", "wohnungsbau",      "Wohnungswesen, Wohnungsbauprämie"),
         ("1036", "raumordnung",      "Raumordnung, Städtebauförderung"),
         ("1037", "gemeinschaftsdienste", "Kommunale Gemeinschaftsdienste")]),
    ("1038", "landwirtschaft", "Ernährung, Landwirtschaft und Forsten", []),
    ("1039", "gewerbe",       "Energie- und Wasserwirtschaft, Gewerbe, Dienstleistungen",
        [("1040", "energie_entsorgung",  "Energie, Wasserversorgung, Entsorgung"),
         ("1041", "abwasser",            "Abwasserentsorgung"),
         ("1042", "abfall",              "Abfallwirtschaft"),
         ("1043", "versicherungswesen",  "Geld- und Versicherungswesen"),
         ("1044", "sonst_gewerbe",       "Sonstiges Gewerbe und Dienstleistungen"),
         ("1045", "reg_foerderung",      "Regionale Fördermaßnahmen")]),
    ("1046", "verkehr",       "Verkehrs- und Nachrichtenwesen",
        [("1047", "strassen",            "Straßen einschl. Verwaltung"),
         ("1048", "autobahnen",          "Bundesautobahnen"),
         ("1049", "bundes_landesstr",    "Bundes- und Landesstraßen"),
         ("1050", "kreisstrassen",       "Kreisstraßen"),
         ("1051", "gemeindestrassen",    "Gemeindestraßen"),
         ("1052", "wasserstrassen",      "Wasserstraßen und Häfen"),
         ("1053", "eisenbahn_oepnv",     "Eisenbahnen und ÖPNV"),
         ("1054", "sonst_verkehr",       "Sonstiges Verkehrswesen")]),
    ("1055", "finanzwirtschaft", "Finanzwirtschaft",
        [("1056", "kapitalvermoegen",    "Grund- und Kapitalvermögen"),
         ("1057", "steuern_finzuw",      "Steuern und Finanzzuweisungen"),
         ("1058", "schulden",            "Schulden"),
         ("1059", "beihilfen",           "Beihilfen und Unterstützungen")]),
    ("1060", "versorgung",    "Versorgung (Beamtenpensionen)", []),
]

OFFSET_BEREINIGTE_LAENDER = 43


def parse_value(v):
    """Convert cell value (Millionen EUR) to integer EUR."""
    if v is None:
        return 0
    if isinstance(v, str):
        v = v.strip()
        if v in ("-", ".", "", "\u2013", "...", "x", "X"):
            return 0
        try:
            v = float(v.replace(",", "."))
        except ValueError:
            return 0
    try:
        return int(round(float(v) * 1_000_000))
    except (ValueError, TypeError):
        return 0


def load_all_data(ws, start_row, year_format):
    """Single-pass read of entire worksheet.
    Returns:
      - aufgaben_rows: dict code -> header_row_number
      - all_data: dict row_number -> list of cell values (indexed by column, 1-based)
    """
    aufgaben_rows = {}
    all_data = {}

    # Collect all codes we need to find
    needed_codes = set(["1001"])
    for code, _, _, sub_codes in TOP_LEVEL_FUNCTIONS:
        needed_codes.add(code)
        for sc, _, _ in sub_codes:
            needed_codes.add(sc)

    # Collect all needed data rows (will be filled after finding headers)
    needed_rows = set()

    print("  Single-pass scan...")
    t0 = time.time()

    # First pass: find all Aufgabenbereich headers
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row,
                            min_col=1, max_col=1, values_only=True):
        # We just need col 1 for detection
        pass

    # Actually we need row numbers. Use non-values_only for first col only.
    row_count = 0
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row,
                            min_col=1, max_col=1, values_only=False):
        row_count += 1
        r = row[0].row
        a = str(row[0].value) if row[0].value else ""

        if year_format == "2017":
            # Fachserie: col A has "NNNN.1" for headers
            if "." in a:
                parts = a.split(".")
                if len(parts) == 2 and parts[1].strip() == "1" and parts[0].strip().isdigit():
                    code = parts[0].strip()
                    if code in needed_codes:
                        aufgaben_rows[code] = r
        else:
            # 2018+ format: col A has "NNNN\n\nNNNN.1"
            if "\n" in a:
                parts = a.split("\n")
                first = parts[0].strip()
                if first and "." not in first and first.isdigit():
                    if first in needed_codes:
                        aufgaben_rows[first] = r

    print("  Found {} Aufgabenbereiche in {} rows ({:.1f}s)".format(
        len(aufgaben_rows), row_count, time.time() - t0))

    # Determine which rows we need data from
    for code, header_row in aufgaben_rows.items():
        data_row = header_row + OFFSET_BEREINIGTE_LAENDER
        needed_rows.add(data_row)

    # Second pass: read only the rows we need (cols 3-19)
    print("  Reading {} data rows...".format(len(needed_rows)))
    t1 = time.time()
    min_needed = min(needed_rows) if needed_rows else start_row
    max_needed = max(needed_rows) if needed_rows else start_row

    for row in ws.iter_rows(min_row=min_needed, max_row=max_needed,
                            min_col=1, max_col=20, values_only=False):
        r = row[0].row
        if r in needed_rows:
            row_data = {}
            for c in row:
                if hasattr(c, "column"):
                    row_data[c.column] = c.value
            all_data[r] = row_data

    print("  Read data in {:.1f}s".format(time.time() - t1))
    return aufgaben_rows, all_data


def extract_laender_data(aufgaben_rows, all_data, code):
    """Extract per-state values for a given code from pre-loaded data."""
    if code not in aufgaben_rows:
        return None
    header_row = aufgaben_rows[code]
    data_row = header_row + OFFSET_BEREINIGTE_LAENDER

    if data_row not in all_data:
        return None

    raw = all_data[data_row]
    result = {"total": parse_value(raw.get(3))}
    for col, (abbrev, _name) in STATE_COLUMNS.items():
        result[abbrev] = parse_value(raw.get(col))
    return result


def build_function_children(aufgaben_rows, all_data, sub_codes, state_abbrev):
    children = []
    for sub_code, sub_id, sub_name in sub_codes:
        data = extract_laender_data(aufgaben_rows, all_data, sub_code)
        if data is None:
            continue
        val = data.get(state_abbrev, 0)
        if val > 0:
            children.append({
                "id": "laender_{}_{}".format(state_abbrev, sub_id),
                "name": sub_name,
                "value": val,
            })
    children.sort(key=lambda x: x["value"], reverse=True)
    return children


def build_state_tree(aufgaben_rows, all_data, state_abbrev, state_name):
    total_data = extract_laender_data(aufgaben_rows, all_data, "1001")
    state_total = total_data.get(state_abbrev, 0) if total_data else 0

    func_children = []
    for code, func_id, func_name, sub_codes in TOP_LEVEL_FUNCTIONS:
        data = extract_laender_data(aufgaben_rows, all_data, code)
        if data is None:
            continue
        val = data.get(state_abbrev, 0)
        if val <= 0:
            continue

        node = {
            "id": "laender_{}_{}".format(state_abbrev, func_id),
            "name": func_name,
            "value": val,
        }

        sub_children = build_function_children(aufgaben_rows, all_data, sub_codes, state_abbrev)
        if sub_children:
            node["children"] = sub_children

        func_children.append(node)

    func_children.sort(key=lambda x: x["value"], reverse=True)

    return {
        "id": "laender_{}".format(state_abbrev),
        "name": state_name,
        "value": state_total,
        "children": func_children,
    }


def process_year(year, cfg):
    print("\n=== Processing {} ===".format(year))
    fpath = os.path.join(DATA_DIR, cfg["file"])
    if not os.path.exists(fpath):
        print("  File not found: {}".format(fpath))
        return None

    t0 = time.time()
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb[cfg["sheet"]]
    print("  Loaded {} in {:.1f}s".format(cfg["file"], time.time() - t0))

    aufgaben_rows, all_data = load_all_data(ws, cfg["start_row"], cfg["format"])

    # Get overall total
    total_data = extract_laender_data(aufgaben_rows, all_data, "1001")
    overall_total = total_data.get("total", 0) if total_data else 0
    print("  Overall Laender total: {:,} EUR".format(overall_total))

    state_children = []
    for col in sorted(STATE_COLUMNS.keys()):
        abbrev, name = STATE_COLUMNS[col]
        state_tree = build_state_tree(aufgaben_rows, all_data, abbrev, name)
        if state_tree["value"] > 0:
            state_children.append(state_tree)
            n_funcs = len(state_tree["children"])
            print("    {} ({}): {:,} EUR, {} functions".format(
                name, abbrev, state_tree["value"], n_funcs))

    state_children.sort(key=lambda x: x["value"], reverse=True)

    tree = {
        "id": "laender",
        "name": "Länder (Bundesländer)",
        "value": overall_total,
        "year": year,
        "source": "Destatis Rechnungsergebnis {}, Bereinigte Ausgaben der Länder".format(year),
        "unit": "EUR",
        "children": state_children,
    }

    wb.close()

    # Save
    outpath = os.path.join(DATA_DIR, "laender_parsed_{}.json".format(year))
    with open(outpath, "w", encoding="utf-8") as f:
        json.dump(tree, f, ensure_ascii=False, indent=2)
    size_kb = os.path.getsize(outpath) / 1024
    print("  Saved: {} ({:.1f} KB)".format(outpath, size_kb))
    print("  {} states total".format(len(tree["children"])))
    print("  Completed in {:.1f}s".format(time.time() - t0))

    return tree


def main():
    print("=" * 70)
    print("Parsing Destatis Rechnungsergebnis 2017 and 2018")
    print("=" * 70)
    sys.stdout.flush()

    configs = {
        2017: {
            "file": "rechnungsergebnis_2017.xlsx",
            "sheet": "2.2",
            "format": "2017",
            "start_row": 4,
        },
        2018: {
            "file": "rechnungsergebnis_2018.xlsx",
            "sheet": "71711-12",
            "format": "2018",
            "start_row": 7,
        },
    }

    for year in sorted(configs.keys()):
        process_year(year, configs[year])
        sys.stdout.flush()

    print("\n" + "=" * 70)
    print("Done.")


if __name__ == "__main__":
    main()
