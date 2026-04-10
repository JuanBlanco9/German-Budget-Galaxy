"""Parse UK OSCAR XLSX files into budget tree JSONs.

Two-pass approach:
  Pass 1: Build 3-level tree (Dept > Org > SubFunction)
  Pass 2: For any SubFunction leaf node that is >5% of its department,
          inject SEGMENT_L4 children for finer granularity.
"""
import json
import os
import re
import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data', 'uk')

FILES = {
    2020: 'BUD_20-21.xlsx',
    2021: 'BUD_21-22.xlsx',
    2022: 'BUD_22-23.xlsx',
    2023: 'BUD_23-24.xlsx',
    2024: 'BUD_24-25.xlsx',
}

# Column indices
COL_DEPT = 6       # DEPARTMENT_GROUP_LONG_NAME
COL_ORG = 8        # ORGANISATION_LONG_NAME
COL_SUBFUNC = 48   # SUB_FUNCTION_LONG_NAME
COL_SEGMENT = 26   # SEGMENT_L4_LONG_NAME
COL_AMOUNT = 89    # AMOUNT

# Threshold: L3 leaf nodes above this % of dept value get SEGMENT_L4 children
SEGMENT_THRESHOLD_PCT = 5.0


def count_nodes(node):
    c = 1
    for ch in node.get('children', []):
        c += count_nodes(ch)
    return c


def max_depth(node, d=0):
    if not node.get('children'):
        return d
    return max(max_depth(ch, d + 1) for ch in node['children'])


def clean_segment_name(raw):
    """Strip OSCAR segment code prefix (e.g. 'X017A049-') and accounting markers."""
    s = re.sub(r'^X\d{3}[A-Z]?\d{3,4}\s*[-–]\s*', '', raw).strip()
    # Remove accounting suffixes: DEL/AME/VOTED/NON-VOTED/DEPT/PROG/ADMIN/NIF/IWC/OWC etc.
    # These appear as trailing markers like "- Voted Dept Ame - Iwc" or "DEL PROG VOTED"
    s = re.sub(r'\s*[-–]\s*(NON[-\s]?VOTED|VOTED)\b.*$', '', s, flags=re.IGNORECASE).strip()
    s = re.sub(r'\s+(DEL|AME|VOTED|NON-VOTED|DEPT|PROG|ADMIN|NON-BUDGET|NON_BUDGET|NIF|CAPITAL)[\s/_A-Z]*$', '', s, flags=re.IGNORECASE).strip()
    # Second pass for remaining markers
    s = re.sub(r'\s+(DEL|AME)\b.*$', '', s, flags=re.IGNORECASE).strip()
    # Capitalize first letter of each word if mostly uppercase
    if s == s.upper() and len(s) > 3:
        s = s.title()
    # Fix common title-case artifacts
    s = s.replace("'S ", "'s ").replace("N.E.C.", "n.e.c.").replace("N.E.c.", "n.e.c.")
    # Remove trailing whitespace and punctuation
    s = s.strip(' -–:')
    return s or raw


def enrich_with_segments(dept_node, departments, segment_data):
    """For L3 leaf nodes >SEGMENT_THRESHOLD_PCT of dept, add SEGMENT_L4 children.

    Called per-department BEFORE flatten, while dept > org > subfunc structure is intact.
    """
    enriched = 0
    dept_name = dept_node['name']
    dept_val = abs(dept_node['value']) or 1

    for org_node in dept_node.get('children', []):
        org_name = org_node['name']

        if org_node.get('children'):
            for prog_node in org_node['children']:
                if prog_node.get('children'):
                    continue
                pct = abs(prog_node['value']) / dept_val * 100
                if pct < SEGMENT_THRESHOLD_PCT:
                    continue

                seg_key = (dept_name, org_name, prog_node['name'])
                segs = segment_data.get(seg_key, {})
                nonzero = {k: v for k, v in segs.items() if round(v) != 0}
                if len(nonzero) < 2:
                    continue

                parent_id = prog_node.get('id', '')
                seg_children = sorted(
                    [{'id': parent_id + '__' + re.sub(r'[^a-z0-9]+', '_', s.lower()).strip('_'),
                      'name': clean_segment_name(s),
                      'value': round(v)}
                     for s, v in nonzero.items()],
                    key=lambda x: -abs(x['value'])
                )
                prog_node['children'] = seg_children
                enriched += 1
        else:
            # org_node is itself a leaf (single-subfunc departments)
            pct = abs(org_node['value']) / dept_val * 100
            if pct < SEGMENT_THRESHOLD_PCT:
                continue
            seg_key = (dept_name, org_name, org_name)
            segs = segment_data.get(seg_key, {})
            nonzero = {k: v for k, v in segs.items() if round(v) != 0}
            if len(nonzero) < 2:
                continue
            parent_id = org_node.get('id', '')
            seg_children = sorted(
                [{'id': parent_id + '__' + re.sub(r'[^a-z0-9]+', '_', s.lower()).strip('_'),
                  'name': clean_segment_name(s),
                  'value': round(v)}
                 for s, v in nonzero.items()],
                key=lambda x: -abs(x['value'])
            )
            org_node['children'] = seg_children
            enriched += 1

    return enriched


for fy, fname in FILES.items():
    out_path = os.path.join(DATA_DIR, f'uk_budget_tree_{fy}.json')

    # Rebuild all years to apply SEGMENT_L4 enrichment
    # (set SKIP_EXISTING=True to only rebuild missing/small files)
    SKIP_EXISTING = False
    if SKIP_EXISTING and os.path.exists(out_path) and os.path.getsize(out_path) > 100000:
        print(f"=== {fy}-{fy+1}: SKIP (exists, {os.path.getsize(out_path)//1024}KB) ===")
        continue

    fpath = os.path.join(DATA_DIR, fname)
    if not os.path.exists(fpath):
        print(f"=== {fy}-{fy+1}: MISSING {fname} ===")
        continue

    print(f"=== {fy}-{fy+1}: Parsing {fname}... ===")

    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    departments = {}
    # Also collect segment data keyed by (dept, org, subfunc) -> {segment: amount}
    segment_data = {}
    row_count = 0

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or len(row) <= COL_AMOUNT:
            continue
        dept = str(row[COL_DEPT] or '').strip()
        org = str(row[COL_ORG] or '').strip()
        subfunc = str(row[COL_SUBFUNC] or '').strip()
        segment = str(row[COL_SEGMENT] or '').strip() if len(row) > COL_SEGMENT else ''
        amount = row[COL_AMOUNT]

        if not dept or amount is None:
            continue
        try:
            amount = float(amount) * 1000  # OSCAR amounts in thousands
        except (ValueError, TypeError):
            continue

        # Clean up department names (remove GROUP suffix)
        dept = dept.replace(' (GROUP)', '').replace(' (group)', '').strip()

        if dept not in departments:
            departments[dept] = {'value': 0, 'orgs': {}}
        departments[dept]['value'] += amount

        if org not in departments[dept]['orgs']:
            departments[dept]['orgs'][org] = {'value': 0, 'progs': {}}
        departments[dept]['orgs'][org]['value'] += amount

        prog = subfunc or 'General'
        if prog not in departments[dept]['orgs'][org]['progs']:
            departments[dept]['orgs'][org]['progs'][prog] = 0
        departments[dept]['orgs'][org]['progs'][prog] += amount

        # Collect segment-level data for enrichment pass
        if segment:
            seg_key = (dept, org, prog)
            if seg_key not in segment_data:
                segment_data[seg_key] = {}
            if segment not in segment_data[seg_key]:
                segment_data[seg_key][segment] = 0
            segment_data[seg_key][segment] += amount

        row_count += 1

    wb.close()
    print(f"  {row_count} rows parsed")

    # Build tree
    children = []
    enriched = 0
    for dept_name, dept_data in departments.items():
        dept_id = re.sub(r'[^a-z0-9]+', '_', dept_name.lower()).strip('_')
        dept_node = {
            'id': dept_id,
            'name': dept_name,
            'value': round(dept_data['value']),
            'children': []
        }
        for org_name, org_data in dept_data['orgs'].items():
            org_id = dept_id + '__' + re.sub(r'[^a-z0-9]+', '_', org_name.lower()).strip('_')
            org_node = {
                'id': org_id,
                'name': org_name,
                'value': round(org_data['value']),
            }
            progs = {k: v for k, v in org_data['progs'].items() if v != 0}
            if len(progs) > 1:
                org_node['children'] = sorted(
                    [{'id': org_id + '__' + re.sub(r'[^a-z0-9]+', '_', p.lower()).strip('_'),
                      'name': p, 'value': round(v)}
                     for p, v in progs.items()],
                    key=lambda x: -abs(x['value'])
                )
            dept_node['children'].append(org_node)

        dept_node['children'].sort(key=lambda x: -abs(x['value']))

        # SEGMENT_L4 enrichment (BEFORE flatten, while dept > org > subfunc structure intact)
        enriched += enrich_with_segments(dept_node, departments, segment_data)

        # Flatten redundant single-org wrapper: if the department has one org (or one
        # dominant org with same name), promote its children directly under the dept.
        if len(dept_node['children']) == 1:
            only = dept_node['children'][0]
            dept_node['children'] = only.get('children', [])
        elif dept_node['children']:
            top = dept_node['children'][0]
            dept_val = abs(dept_node['value']) or 1
            if (top['name'].lower() == dept_name.lower()
                    and abs(top['value']) / dept_val > 0.90
                    and top.get('children')):
                # Promote dominant same-name org's children; keep smaller orgs as siblings
                promoted = top.get('children', [])
                siblings = dept_node['children'][1:]
                dept_node['children'] = promoted + siblings
        children.append(dept_node)

    children.sort(key=lambda x: -abs(x['value']))
    children = [c for c in children if abs(c['value']) > 100000]

    total = sum(c['value'] for c in children)
    fy_label = f'{fy}-{str(fy+1)[2:]}'
    tree = {
        'name': f'UK Government Spending {fy_label}',
        'value': total,
        'children': children
    }

    n_after = count_nodes(tree)
    d = max_depth(tree)
    print(f"  GBP{total/1e9:.0f}B | {len(children)} depts | {n_after} nodes | {d} levels | {enriched} nodes enriched")

    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(tree, f)
    print(f"  -> {out_path} ({os.path.getsize(out_path)//1024}KB)\n")
