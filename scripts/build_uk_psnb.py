#!/usr/bin/env python3
"""
Extract UK Public Sector Net Borrowing (PSNB) + Debt (PSND) historical series
from the OBR Historical Public Finances Database XLSX.

Output: data/uk/fiscal/uk_psnb_historical.json
"""
import hashlib
import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'data' / 'uk' / 'fiscal' / 'obr_historical_public_finances.xlsx'
OUT = ROOT / 'data' / 'uk' / 'fiscal' / 'uk_psnb_historical.json'


def sha256(p: Path) -> str:
    h = hashlib.sha256()
    with open(p, 'rb') as f:
        for c in iter(lambda: f.read(1 << 20), b''):
            h.update(c)
    return h.hexdigest()


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb['Aggregates (£m)']
    rows = list(ws.iter_rows(values_only=True))

    data = []
    for row in rows[8:]:
        yr = row[1]
        if not isinstance(yr, str):
            continue
        try:
            start = int(yr.split('-')[0])
        except ValueError:
            continue
        if start < 2000 or start > 2030:
            continue
        psnb = float(row[2]) if isinstance(row[2], (int, float)) else None
        psnd = float(row[3]) if isinstance(row[3], (int, float)) else None
        gdp = float(row[4]) if isinstance(row[4], (int, float)) else None
        data.append({
            'fiscal_year_label': yr,
            'fiscal_year_start': start,
            'psnb_gbp_m': round(psnb, 0) if psnb is not None else None,
            'psnd_gbp_m': round(psnd, 0) if psnd is not None else None,
            'nominal_gdp_gbp_m': round(gdp, 0) if gdp is not None else None,
            'psnb_pct_of_gdp': (
                round(psnb / gdp * 100, 2)
                if (psnb is not None and gdp) else None
            ),
            'psnd_pct_of_gdp': (
                round(psnd / gdp * 100, 2)
                if (psnd is not None and gdp) else None
            ),
        })

    file_path_rel = str(XLSX.relative_to(ROOT)).replace('\\', '/')

    out = {
        'country': 'uk',
        'subject': 'public_sector_borrowing_and_debt',
        'unit': 'GBP_million',
        'source': {
            'publisher': 'Office for Budget Responsibility (OBR)',
            'dataset': 'Historical Public Finances Database',
            'landing_url': 'https://obr.uk/data/',
            'file_url': (
                'https://obr.uk/download/historical-public-finances-database/'
            ),
            'file_path': file_path_rel,
            'file_sha256': sha256(XLSX),
            'downloaded_at': date.today().isoformat(),
            'sheet': 'Aggregates (£m)',
        },
        'note': (
            'PSNB = Public Sector Net Borrowing (annual deficit; positive = gov '
            'spent more than it received). PSND = Public Sector Net Debt '
            '(cumulative stock). GDP is nominal. Values from 2000-01 onwards; '
            'full series in source goes back to 1700-01.'
        ),
        'series': data,
    }

    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    print(f'Wrote {OUT.relative_to(ROOT)}')
    print(f'Years: {len(data)}')
    print('Last 8:')
    for d in data[-8:]:
        print(
            f"  {d['fiscal_year_label']}  "
            f"PSNB £{d['psnb_gbp_m']:>8,.0f}m ({d['psnb_pct_of_gdp']:>5.1f}% GDP)  "
            f"GDP £{d['nominal_gdp_gbp_m']:>10,.0f}m"
        )


if __name__ == '__main__':
    main()
