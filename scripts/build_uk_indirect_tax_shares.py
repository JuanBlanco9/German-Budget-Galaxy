#!/usr/bin/env python3
"""
Parse ONS 'Effects of Taxes and Benefits' historical XLSX (2017-18 sheet)
into per-decile indirect-tax shares.

These shares let us estimate VAT + other indirect taxes for any user given
their disposable income, by picking the nearest decile.

The ONS dataset we use here is the most recent available
`incometaxandbenefitdatabyincomedecileforallhouseholds.xlsx`, last
updated May 2019. Decile RATIOS (VAT as % of disposable income per decile)
are stable year-to-year, so using these as a distribution is defensible;
scaling by the year's HMRC headline totals keeps the absolute figures current.

Output: data/uk/fiscal/uk_indirect_tax_shares_by_decile.json
"""
import hashlib
import json
from pathlib import Path
from datetime import date

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'data' / 'uk' / 'fiscal' / 'ons_etb_by_decile_all.xlsx'
OUT = ROOT / 'data' / 'uk' / 'fiscal' / 'uk_indirect_tax_shares_by_decile.json'

TARGET_SHEET = '2017-18'

# Row indices (0-based) inside the 2017-18 sheet
ROWS = {
    'decile_upper_equiv_gross': 6,
    'n_households': 8,
    'gross_income': 46,
    'equivalised_gross_income': 48,
    'income_tax': 51,
    'employee_ni': 52,
    'council_tax': 53,
    'ct_less_rebates': 54,
    'direct_taxes_total': 55,
    'disposable_income': 57,
    'equivalised_disposable_income': 59,
    'vat': 63,
    'tobacco_duty': 64,
    'beer_cider_duty': 65,
    'wines_spirits_duty': 66,
    'fuel_duty': 67,
    'ved': 68,
    'tv_license': 69,
    'sdlt': 70,
    'customs': 71,
    'betting': 72,
    'ipt': 73,
    'apd': 74,
    'national_lottery': 75,
    'other_indirect': 76,
    'commercial_industrial_rates': 79,
    'employer_ni': 80,
}


def sha256(p: Path) -> str:
    h = hashlib.sha256()
    with open(p, 'rb') as f:
        for c in iter(lambda: f.read(1 << 20), b''):
            h.update(c)
    return h.hexdigest()


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[TARGET_SHEET]
    rows = list(ws.iter_rows(values_only=True))

    def decile_values(row_idx):
        r = rows[row_idx]
        # Columns B-K (1..10) are deciles 1-10; column L (11) is "All"
        return [float(r[i]) if isinstance(r[i], (int, float)) else None for i in range(1, 12)]

    data = {k: decile_values(i) for k, i in ROWS.items()}

    # Build per-decile record. Decile 1 = Bottom, 10 = Top, index 11 = All.
    deciles = []
    for d in range(10):
        disp = data['disposable_income'][d]
        eq_disp = data['equivalised_disposable_income'][d]
        vat = data['vat'][d]
        indirect_total = sum(
            v[d] or 0 for v in [
                data['vat'], data['tobacco_duty'], data['beer_cider_duty'],
                data['wines_spirits_duty'], data['fuel_duty'], data['ved'],
                data['tv_license'], data['sdlt'], data['customs'],
                data['betting'], data['ipt'], data['apd'], data['national_lottery'],
                data['other_indirect'],
            ] if v[d] is not None
        )

        deciles.append({
            'decile': d + 1,
            'label': ['Bottom', '2nd', '3rd', '4th', '5th',
                      '6th', '7th', '8th', '9th', 'Top'][d],
            'equivalised_disposable_income_gbp': eq_disp,
            'disposable_income_gbp': disp,
            'gross_income_gbp': data['gross_income'][d],
            'indirect_taxes_gbp': {
                'vat': data['vat'][d],
                'tobacco_duty': data['tobacco_duty'][d],
                'beer_cider_duty': data['beer_cider_duty'][d],
                'wines_spirits_duty': data['wines_spirits_duty'][d],
                'fuel_duty': data['fuel_duty'][d],
                'ved': data['ved'][d],
                'tv_license': data['tv_license'][d],
                'sdlt': data['sdlt'][d],
                'customs': data['customs'][d],
                'betting': data['betting'][d],
                'insurance_premium_tax': data['ipt'][d],
                'air_passenger_duty': data['apd'][d],
                'national_lottery': data['national_lottery'][d],
                'other': data['other_indirect'][d],
            },
            'total_indirect_taxes_gbp': round(indirect_total, 0),
            'ratios_of_disposable_income': {
                'vat': round(vat / disp, 4) if disp else None,
                'total_indirect': round(indirect_total / disp, 4) if disp else None,
                'fuel_duty': round(data['fuel_duty'][d] / disp, 4) if disp else None,
                'alcohol': round(
                    ((data['beer_cider_duty'][d] or 0) + (data['wines_spirits_duty'][d] or 0))
                    / disp, 4) if disp else None,
                'tobacco': round(data['tobacco_duty'][d] / disp, 4) if disp else None,
            },
        })

    # Decile thresholds on equivalised gross income (the published cutoffs)
    decile_cutoffs = data['decile_upper_equiv_gross']
    # Cutoffs only start at decile 2 (row shows upper bound of prev decile). Normalise:
    cutoffs_clean = []
    for d_idx, v in enumerate(decile_cutoffs[:10]):
        if v is not None:
            cutoffs_clean.append({'decile': d_idx, 'upper_equivalised_gross_gbp': v})

    out = {
        'country': 'uk',
        'subject': 'indirect_tax_shares_by_income_decile',
        'ons_fiscal_year_source': '2017-18',
        'unit': 'GBP_per_year_per_household',
        'note': (
            'Decile ratios (indirect tax as share of disposable income) from the '
            'ONS "Effects of Taxes and Benefits" historical dataset. Ratios are '
            'stable year-to-year and serve as the distribution basis. For absolute '
            'current-year estimates, scale by the ratio of (current VAT/duty take) '
            'to (2017-18 VAT/duty take), or simply use the ratios directly against '
            'the user\'s current disposable income. Decile cutoffs are on equivalised '
            'gross income; a single-person household has equivalisation factor 1.0, '
            'a couple ~1.5, each child +0.3.'
        ),
        'source': {
            'publisher': 'Office for National Statistics (ONS)',
            'dataset': (
                'Effects of taxes and benefits on UK household income — '
                'Income, tax and benefit data by income decile for all households '
                '(historical dataset)'
            ),
            'landing_url': (
                'https://www.ons.gov.uk/peoplepopulationandcommunity/'
                'personalandhouseholdfinances/incomeandwealth/datasets/'
                'theeffectsoftaxesandbenefitsonhouseholdincomehistoricaldatasets'
            ),
            'file_url': (
                'https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/'
                'personalandhouseholdfinances/incomeandwealth/datasets/'
                'theeffectsoftaxesandbenefitsonhouseholdincomehistoricaldatasets/'
                'incometaxandbenefitdatabyincomedecileforallhouseholds/'
                'incometaxandbenefitdatabyincomedecileforallhouseholds.xlsx'
            ),
            'file_sha256': sha256(XLSX),
            'downloaded_at': date.today().isoformat(),
            'sheet': TARGET_SHEET,
        },
        'decile_cutoffs_equivalised_gross_income_gbp': cutoffs_clean,
        'deciles': deciles,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    print(f'Wrote {OUT.relative_to(ROOT)}')
    print('\nVAT share of disposable income by decile:')
    for d in deciles:
        r = d['ratios_of_disposable_income']['vat']
        print(f'  D{d["decile"]:2d} ({d["label"]:>6s}): {r*100:5.2f}% — '
              f'£{d["indirect_taxes_gbp"]["vat"]:,.0f} on £{d["disposable_income_gbp"]:,.0f}')


if __name__ == '__main__':
    main()
